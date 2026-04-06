import pandas as pd
from typing import List
from core.models import DadosCondominio, ValidationStatus
from loguru import logger
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


class ExcelExporter:
    """Consolida os dados processados e gera o Excel final com 3 abas e formatação."""

    def __init__(self, output_dir: str, filename: str):
        self.output_path = os.path.join(output_dir, filename)
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

    def export(self, resultados: List[DadosCondominio]):
        """Cria o arquivo Excel com as abas detalhadas e resumidas."""
        if not resultados:
            logger.warning("Nenhum resultado para exportar.")
            return

        # 1. Aba 'Detalhado' — uma linha por apartamento, despesas como colunas
        apt_rows = []
        despesa_rows = []

        raw_rows = []  # aba Despesas_Raw: um item por linha com tipo e volume

        for cond in resultados:
            for bloco in cond.blocos:
                for apt in bloco.apartamentos:
                    apt_rows.append({
                        "Arquivo_Origem":     cond.arquivo_origem,
                        "Modelo_PDF":         cond.modelo_pdf,
                        "Condominio":         cond.condominio,
                        "Bloco":              bloco.bloco,
                        "Apartamento":        apt.apartamento,
                        "Total_Impresso_PDF": apt.total_impresso,
                        "Soma_Calculada":     apt.soma_calculada,
                        "Taxa_Assertividade": apt.taxa_assertividade,
                        "Status_Validacao":   apt.status_validacao.value,
                        "Camada_Extracao":    cond.camada_extracao.value,
                        "Pagina_Origem":      apt.pagina_origem,
                    })
                    for item in apt.despesas:
                        despesa_rows.append({
                            "Arquivo_Origem":    cond.arquivo_origem,
                            "Bloco":             bloco.bloco,
                            "Apartamento":       apt.apartamento,
                            "Descricao_Despesa": item.descricao,
                            "Valor_Despesa":     item.valor,
                        })
                        raw_rows.append({
                            "Arquivo_Origem":    cond.arquivo_origem,
                            "Condominio":        cond.condominio,
                            "Bloco":             bloco.bloco,
                            "Apartamento":       apt.apartamento,
                            "Descricao_Despesa": item.descricao,
                            "Tipo_Despesa":      item.tipo_despesa or "NAO_CLASSIFICADO",
                            "Valor_Despesa":     item.valor,
                            "Volume_M3":         item.volume_m3,
                            "Pagina_Origem":     apt.pagina_origem,
                        })

        df_apt = pd.DataFrame(apt_rows)

        if despesa_rows:
            df_despesas_raw = pd.DataFrame(despesa_rows)
            df_pivot = (
                df_despesas_raw
                .pivot_table(
                    index=["Arquivo_Origem", "Bloco", "Apartamento"],
                    columns="Descricao_Despesa",
                    values="Valor_Despesa",
                    aggfunc="sum",
                )
                .fillna(0)
                .reset_index()
            )
            df_pivot.columns.name = None
            df_detalhado = df_apt.merge(
                df_pivot,
                on=["Arquivo_Origem", "Bloco", "Apartamento"],
                how="left",
            )
        else:
            df_detalhado = df_apt

        # Garantir 0 em colunas de despesa que ficaram NaN após o merge
        cols_meta = [
            "Arquivo_Origem", "Modelo_PDF", "Condominio", "Bloco", "Apartamento",
            "Total_Impresso_PDF", "Soma_Calculada", "Taxa_Assertividade",
            "Status_Validacao", "Camada_Extracao", "Pagina_Origem",
        ]
        cols_despesa = [c for c in df_detalhado.columns if c not in cols_meta]
        df_detalhado[cols_despesa] = df_detalhado[cols_despesa].fillna(0)

        # 2. Aba 'Resumo_PDF'
        resumo_pdf_rows = []
        for cond in resultados:
            total_pdf = sum(a.total_impresso or 0 for b in cond.blocos for a in b.apartamentos)
            total_ext = sum(a.soma_calculada for b in cond.blocos for a in b.apartamentos)
            divergentes = sum(
                1 for b in cond.blocos for a in b.apartamentos
                if a.status_validacao == ValidationStatus.DIVERGENTE
            )
            resumo_pdf_rows.append({
                "Arquivo":         cond.arquivo_origem,
                "Modelo":          cond.modelo_pdf,
                "Total_Relatório": total_pdf,
                "Total_Extraído":  total_ext,
                "Diferença":       abs(total_pdf - total_ext),
                "Acurácia_%":      (total_ext / total_pdf * 100) if total_pdf > 0 else 0,
                "PDFs_Divergentes": divergentes,
            })
        df_resumo_pdf = pd.DataFrame(resumo_pdf_rows)

        # 3. Aba 'Resumo_Bloco'
        resumo_bloco_rows = []
        for cond in resultados:
            for bloco in cond.blocos:
                total_bloco_pdf = sum(a.total_impresso or 0 for a in bloco.apartamentos)
                total_bloco_ext = sum(a.soma_calculada for a in bloco.apartamentos)
                status = "OK" if abs(total_bloco_pdf - total_bloco_ext) < 1.0 else "DIVERGENTE"
                resumo_bloco_rows.append({
                    "Arquivo":             cond.arquivo_origem,
                    "Bloco":               bloco.bloco,
                    "Total_Bloco_PDF":     total_bloco_pdf,
                    "Total_Bloco_Extraído": total_bloco_ext,
                    "Status":              status,
                })
        df_resumo_bloco = pd.DataFrame(resumo_bloco_rows)

        df_raw = pd.DataFrame(raw_rows) if raw_rows else pd.DataFrame()

        # Gravar as 4 abas
        with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
            df_detalhado.to_excel(writer, sheet_name='Detalhado', index=False)
            df_resumo_pdf.to_excel(writer, sheet_name='Resumo_PDF', index=False)
            df_resumo_bloco.to_excel(writer, sheet_name='Resumo_Bloco', index=False)
            if not df_raw.empty:
                df_raw.to_excel(writer, sheet_name='Despesas_Raw', index=False)

        self._aplicar_formatacao()
        logger.success(f"Arquivo consolidado gerado com sucesso: {self.output_path}")

    def _aplicar_formatacao(self):
        """Aplica cor vermelha nas linhas com Status_Validacao = DIVERGENTE."""
        wb = load_workbook(self.output_path)
        ws = wb['Detalhado']
        fill_red = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

        status_col_idx = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == "Status_Validacao":
                status_col_idx = idx
                break

        if status_col_idx:
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=status_col_idx).value == "DIVERGENTE":
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = fill_red

        wb.save(self.output_path)
