import pandas as pd
from typing import List
from models import DadosCondominio, ValidationStatus
from loguru import logger
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class ExcelExporter:
    """Consolida os dados processados e gera o Excel final com 3 abas e formatação."""

    def __init__(self, output_dir: str, filename: str):
        self.output_path = os.path.join(output_dir, filename)
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

    def export(self, resultados: List[DadosCondominio]):
        """Cria o arquivo Excel com as abas detalhadas e resumidas."""
        if not resultados:
            logger.warning("Nenhum resultado para exportar.")
            return

        # 1. Preparar Dados para a Aba 'Detalhado'
        detalhado_rows = []
        for cond in resultados:
            for bloco in cond.blocos:
                for apt in bloco.apartamentos:
                    for item in apt.despesas:
                        detalhado_rows.append({
                            "Arquivo_Origem": cond.arquivo_origem,
                            "Modelo_PDF": cond.modelo_pdf,
                            "Condominio": cond.condominio,
                            "Bloco": bloco.bloco,
                            "Apartamento": apt.apartamento,
                            "Descricao_Despesa": item.descricao,
                            "Valor_Despesa": item.valor,
                            "Total_Impresso_PDF": apt.total_impresso,
                            "Soma_Calculada": apt.soma_calculada,
                            "Taxa_Assertividade": apt.taxa_assertividade,
                            "Status_Validacao": apt.status_validacao.value,
                            "Camada_Extracao": cond.camada_extracao.value,
                            "Pagina_Origem": apt.pagina_origem
                        })
        df_detalhado = pd.DataFrame(detalhado_rows)

        # 2. Preparar Dados para a Aba 'Resumo_PDF'
        resumo_pdf_rows = []
        for cond in resultados:
            total_pdf = sum(a.total_impresso or 0 for b in cond.blocos for a in b.apartamentos)
            total_ext = sum(a.soma_calculada for b in cond.blocos for a in b.apartamentos)
            divergentes = sum(1 for b in cond.blocos for a in b.apartamentos if a.status_validacao == ValidationStatus.DIVERGENTE)
            
            resumo_pdf_rows.append({
                "Arquivo": cond.arquivo_origem,
                "Modelo": cond.modelo_pdf,
                "Total_Relatório": total_pdf,
                "Total_Extraído": total_ext,
                "Diferença": abs(total_pdf - total_ext),
                "Acurácia_%": (total_ext / total_pdf * 100) if total_pdf > 0 else 0,
                "PDFs_Divergentes": divergentes
            })
        df_resumo_pdf = pd.DataFrame(resumo_pdf_rows)

        # 3. Preparar Dados para a Aba 'Resumo_Bloco'
        resumo_bloco_rows = []
        for cond in resultados:
            for bloco in cond.blocos:
                total_bloco_pdf = sum(a.total_impresso or 0 for a in bloco.apartamentos)
                total_bloco_ext = sum(a.soma_calculada for a in bloco.apartamentos)
                status = "OK" if abs(total_bloco_pdf - total_bloco_ext) < 1.0 else "DIVERGENTE"
                
                resumo_bloco_rows.append({
                    "Arquivo": cond.arquivo_origem,
                    "Bloco": bloco.bloco,
                    "Total_Bloco_PDF": total_bloco_pdf,
                    "Total_Bloco_Extraído": total_bloco_ext,
                    "Status": status
                })
        df_resumo_bloco = pd.DataFrame(resumo_bloco_rows)

        # Gravar as 3 abas
        with pd.ExcelWriter(self.output_path, engine='openpyxl') as writer:
            df_detalhado.to_excel(writer, sheet_name='Detalhado', index=False)
            df_resumo_pdf.to_excel(writer, sheet_name='Resumo_PDF', index=False)
            df_resumo_bloco.to_excel(writer, sheet_name='Resumo_Bloco', index=False)

        self._aplicar_formatacao()
        logger.success(f"Arquivo consolidado gerado com sucesso: {self.output_path}")

    def _aplicar_formatacao(self):
        """Aplica cores vermelhas nas células divergentes (openpyxl)."""
        wb = load_workbook(self.output_path)
        ws = wb['Detalhado']
        fill_red = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        
        # Encontrar índice da coluna 'Status_Validacao'
        status_col_idx = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == "Status_Validacao":
                status_col_idx = idx
                break
        
        if status_col_idx:
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=status_col_idx).value == "DIVERGENTE":
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = fill_red
        
        wb.save(self.output_path)
